from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# 엑셀파일에서 이메일 주소 읽어와 메일 보내기
# 네이버/구글 일 1000건 이상시 제한됨 주의

load_wb = load_workbook("이메일주소.xlsx", data_only=True)
load_ws =load_wb.active

for i in range(1,load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:",recv_email_value)
    try:
        send_email = "네이버계정 메일주소"
        send_pwd = "네이버 비밀번호"

        recv_email = recv_email_value

        smtp_name = "smtp.naver.com" 
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] ="엑셀에서 메일주소를 읽어 자동으로 보내는 메일입니다."
        msg['From'] = send_email 
        msg['To'] = recv_email 

        text =  """
                메일내용 입니다.
                감사합니다.
                """

        msg.attach( MIMEText(text) ) 

        s=smtplib.SMTP( smtp_name , smtp_port )
        s.starttls()
        s.login( send_email , send_pwd )
        s.sendmail( send_email, recv_email, msg.as_string() )
        s.quit()
    except:
        print("에러:",recv_email_value)