import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

test_string = """
aaa@bbb.com
123@abc.co.kr
test@hello.kr
ok@ok.co.kr
ok@ok.co.kr
ok@ok.co.kr
no.co.kr
no.kr
"""

# 이메일 정보 정규식 으로 취합
results = re.findall(r'[\w\.-]+@[\w\.-]+', test_string)
print(results)
print("===============================================")

# set 중복 제거 처리를 통해 unique 이메일 주소 취합
results = list(set(results))
print(results)
print("===============================================")


# 웹 사이트에서 이메일 주소 취득후 엑셀 파일로 저장
url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
        }

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))
print(results)

try:
    wb = load_workbook("email.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save("email.xlsx")
