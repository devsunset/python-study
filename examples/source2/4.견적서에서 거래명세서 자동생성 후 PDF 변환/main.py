from openpyxl import load_workbook

견적서_파일경로 = "견적서_샘플.xlsx"

견적서_wb = load_workbook(견적서_파일경로, data_only=True)
견적서_ws = 견적서_wb.active

#하나의 셀 읽기
견적받는자 = 견적서_ws['A4'].value
소계 = 견적서_ws['X25'].value
부가세 =  견적서_ws['X26'].value
총합계금액 = 견적서_ws['X27'].value

print("견적받는자:",견적받는자)
print("소계:",소계)
print("부가세:",부가세)
print("총합계금액:",총합계금액)

#여러개의 셀 반복하여 읽기
품목명_list = []
수량_list = []
단가_list = []
금액_list = []
for data in 견적서_ws['C13':'X24']:
    for cell in data:
        if cell.column == ord('C')-64:
            if cell.value is not None:
                품목명_list.append(cell.value)
        if cell.column == ord('R')-64:
            if cell.value is not None:
                수량_list.append(cell.value)
        if cell.column == ord('T')-64:
            if cell.value is not None:
                단가_list.append(cell.value)
        if cell.column == ord('X')-64:
            if cell.value is not None:
                금액_list.append(cell.value)

print("품목명:",품목명_list)
print("수량:",수량_list)
print("단가:",단가_list)
print("금액:",금액_list)


from datetime import date

거래명세표_파일경로 = "거래명세표_샘플.xlsx"

거래명세표_wb = load_workbook(거래명세표_파일경로, data_only=False)
거래명세표_ws = 거래명세표_wb.active

today = date.today()
거래명세표_ws['C4'].value = today.year
거래명세표_ws['E4'].value = today.month
거래명세표_ws['G4'].value = today.day

거래명세표_ws['C6'].value = 견적받는자

for i in range(len(품목명_list)):
    거래명세표_ws.cell(row=i+10, column=ord('B')-64).value = 품목명_list[i]
    거래명세표_ws.cell(row=i+10, column=ord('G')-64).value = 수량_list[i]
    거래명세표_ws.cell(row=i+10, column=ord('I')-64).value = 단가_list[i]

거래명세표_wb.save("거래명세표_" + 견적받는자 +".xlsx")


import win32com.client
import os

excel = win32com.client.Dispatch("Excel.Application")

load_file_path = "거래명세표_땡땡초등학교.xlsx"

wb = excel.Workbooks.Open(load_file_path)
ws_sheet = wb.ActiveSheet
ws_sheet.Select()

save_path = os.path.splitext(load_file_path)[0] + '.pdf'
wb.ActiveSheet.ExportAsFixedFormat(0, save_path)

wb.Close(False)
excel.Quit()
