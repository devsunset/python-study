import glob
from openpyxl import load_workbook

발주서들 = glob.glob(r'2.여러개의 엑셀 발주수량 합치기\발주서_*.xlsx')
print(발주서들)

발주처_list =[]
for 발주서 in 발주서들:
    wb = load_workbook(발주서, data_only=True)
    ws = wb.active
    발주처_list.append(ws['B1'].value)

print("발주처:",발주처_list)