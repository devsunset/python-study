import glob
from openpyxl import load_workbook
from openpyxl import Workbook

판매보고들 = glob.glob(r'1.여러개의 엑셀 판매보고서 합치기\판매보고_*.xlsx')

판매점_list =[]
날짜_list = []
금액_list = []
for 판매보고 in 판매보고들:
    wb = load_workbook(판매보고, data_only=True)
    ws = wb.active
    판매점_list.append(ws['B1'].value)
    날짜_list.append(str(ws['B2'].value))
    금액_list.append(ws['B3'].value)

print(판매점_list)
print(날짜_list)
print(금액_list)


try:
    wb = load_workbook(r"1.여러개의 엑셀 판매보고서 합치기\결과.xlsx", data_only=True)
    ws  = wb.active
except:
    wb = Workbook()
    ws = wb.active

for i in range(len(판매점_list)):
    ws.cell(row=i+1,column=1).value = 판매점_list[i]
    ws.cell(row=i+1,column=2).value = 날짜_list[i]
    ws.cell(row=i+1,column=3).value = 금액_list[i]

wb.save(r"1.여러개의 엑셀 판매보고서 합치기\결과.xlsx")