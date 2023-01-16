import glob
from openpyxl import load_workbook
import pandas as pd

발주서들 = glob.glob(r'2.여러개의 엑셀 발주수량 합치기\발주서_*.xlsx')
print(발주서들)

발주처_list =[]
물품_list = []
수량_list = []
for 발주서 in 발주서들:
    wb = load_workbook(발주서, data_only=True)
    ws = wb.active
    발주처_list.append(ws['B1'].value)
    for data in ws['A4':'B12']:
        for cell in data:
            if cell.column == 1:
                if cell.value is not None:
                    물품_list.append(cell.value)
            elif cell.column == 2:
                if cell.value is not None:
                    수량_list.append(cell.value)

print("발주처:",발주처_list)
print("물품:",물품_list)
print("수량:",수량_list)


df = pd.DataFrame({ '물품' :  물품_list,
                    '수량' :  수량_list
                    })

df = df.groupby('물품').sum()

df.to_excel(r'2.여러개의 엑셀 발주수량 합치기\통합.xlsx')