from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import requests
import re
import datetime

import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

#코로나확진자 데이터 얻기
url = "http://openapi.data.go.kr/openapi/service/rest/Covid19/getCovid19SidoInfStateJson?serviceKey="
api_key = "fFWLxGIoKo8cQCIuS5Is1fVoiKXkdls%2FU5DSGRwzmbiwIBI0nlz5V6jllexlrGLKR9y8wV3E3i0SMPTLtAhyvw%3D%3D"

def get_today_covid19(url,api_key):
    now = datetime.datetime.now()
    yyyymmdd = now.strftime('%Y%m%d')
    
    page_no =  "&pageNo=1&numOfRows=30&"
    today = "startCreateDt=" + yyyymmdd + "&endCreateDt=" + yyyymmdd

    response = requests.get(url + api_key + page_no + today)

    gubun = re.findall(r'<gubun>(.+?)</gubun>',response.text)
    incDec = re.findall(r'<incDec>(.+?)</incDec>',response.text)

    return yyyymmdd,gubun,incDec

날짜, 지역, 확진자수 = get_today_covid19(url,api_key)


#엑셀파일 읽기 or 생성
try:
    wb = load_workbook("코로나확진자보고서.xlsx", data_only=True)
    ws  = wb.active
except:
    wb = Workbook()
    ws = wb.active

#코로나 현황 데이터 쓰기
ws.cell(row=1,column=1).value = 날짜 + "일 코로나 현황"
ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=len(확진자수))
ws.cell(row=1,column=1).alignment = Alignment(horizontal="center", vertical="center")

for i in range(len(확진자수)):
    ws.cell(row=2,column=i+1).value = 지역[i]
    ws.cell(row=3,column=i+1).value = int(확진자수[i])

#차트그리기
data = Reference(ws, min_row=2, max_row=3, min_col=1, max_col=18)
cats = Reference(ws, min_row=1, max_row=2, min_col=2, max_col=18)

chart = BarChart()
chart.height = 10
chart.width = 30
chart.legend = None
chart.add_data(data, titles_from_data=True, from_rows=True)
chart.set_categories(cats)

ws.add_chart(chart, "A5")

#저장
wb.save("코로나확진자보고서.xlsx")
